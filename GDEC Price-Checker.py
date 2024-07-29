import os
import re
import openpyxl
import pandas as pd
from difflib import get_close_matches
import customtkinter as ctk
from tkinter import Label, filedialog, messagebox, StringVar, BooleanVar
import webbrowser
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font


def open_file_directory(file_path):
    webbrowser.open(f'file:///{os.path.dirname(file_path)}')


def create_button(tab, text, command):
    return ctk.CTkButton(tab, text=text, command=command)


def create_tab(tabview, tab_name, regular_function, manual_function, promo_sheet_name, window_width):
    tab = tabview.add(tab_name)

    tb_file_path = StringVar()
    pwp_file_path = StringVar()
    save_dir = StringVar()
    selected_promo = StringVar(value="Required Field")
    is_manual = BooleanVar(value=False)

    expected_headers = {
        "Lazada": {
            "manual": ["Shop SKU", "SpecialPrice", "SpecialPrice Start", "SpecialPrice End"],
            "regular": ["Shop Sku", "Campaign Price", "Recommended Price"]
        },
        "Shopee": {
            "manual": ["Variation ID", "Product ID", "Discount price"],
            "regular": ["Variation ID", "Recommended Campaign Price", "Campaign Price"]
        },
        "TikTok": {
            "manual": ["Product_id (required)", "SKU_id (required)", "Deal Price (required)"],
            "regular": ["SKU ID", "Product ID", "Campaign price"]
        }
    }

    def is_file_open(file_path):
        try:
            with open(file_path, 'r+'):
                return False
        except IOError:
            return True

    def select_tb_file():
        file_path = filedialog.askopenfilename(title="Select the TB file", filetypes=[("Excel files", "*.xlsx")])
        platform_headers = expected_headers[tab_name]
        if validate_tb_file(file_path, platform_headers):
            tb_file_path.set(file_path)
            tb_label.configure(text=f"{os.path.basename(file_path)}")
        else:
            messagebox.showerror("Error", f"Please select a {tab_name} TB file.")

    def validate_tb_file(file_path, platform_headers):
        try:
            if tab_name == "TikTok" and not is_manual.get():
                tb_df = pd.read_excel(file_path, sheet_name=0, header=1)
            else:
                tb_df = pd.read_excel(file_path, sheet_name=0)
            tb_headers = list(tb_df.columns)
            if is_manual.get():
                return set(platform_headers["manual"]).issubset(tb_headers)
            else:
                return set(platform_headers["regular"]).issubset(tb_headers)
        except Exception as e:
            return False

    def select_pwp_file():
        file_path = filedialog.askopenfilename(title="Select the PWP file", filetypes=[("Excel files", "*.xlsx")])
        try:
            xl = pd.ExcelFile(file_path)
            if promo_sheet_name not in xl.sheet_names:
                messagebox.showerror("Error", "Please select the correct PWP file")
                return
        except Exception as e:
            messagebox.showerror("Error", "Please select the PWP file first")
            return

        pwp_file_path.set(file_path)
        pwp_label.configure(text=f"{os.path.basename(file_path)}")
        populate_promo_dropdown()
        promo_dropdown.configure(state="normal")

    def populate_promo_dropdown():
        if not pwp_file_path.get():
            messagebox.showerror("Error", "Please select the PWP file first.")
            return

        pwp_df = pd.read_excel(pwp_file_path.get(), sheet_name=promo_sheet_name, header=None)
        pwp_df.columns = pwp_df.iloc[5]
        pwp_df = pwp_df.drop(5)

        promo_name_col = 'Promo Name (Scheme)'
        promo_names = pwp_df[promo_name_col].dropna().unique()

        # Convert all promo names to strings and drop NaN values
        promo_names = [str(name) for name in promo_names if pd.notna(name)]

        if promo_names:
            selected_promo.set(promo_names[0])
            promo_dropdown.set(promo_names[0])
            promo_dropdown.configure(values=promo_names)
        else:
            messagebox.showerror("Error", "No valid promo names found in the PWP file.")

    def select_save_dir():
        directory = filedialog.askdirectory(title="Select the directory to save the updated file")
        save_dir.set(directory)
        save_dir_label.configure(text=f"{directory}")

    def process_files():
        if not tb_file_path.get() or not pwp_file_path.get() or not save_dir.get() or not selected_promo.get() or selected_promo.get() == "Required Field":
            messagebox.showerror("Error", "Please select all files, directories, and promo.")
            return

        if is_file_open(tb_file_path.get()):
            messagebox.showerror("Error",
                                 f"TB file is open: {os.path.basename(tb_file_path.get())}. Please close it before proceeding.")
            return
        if is_file_open(pwp_file_path.get()):
            messagebox.showerror("Error",
                                 f"PWP file is open: {os.path.basename(pwp_file_path.get())}. Please close it before proceeding.")
            return

        updated_tb_file_path = os.path.join(save_dir.get(), "Updated_" + os.path.basename(tb_file_path.get()))
        if os.path.exists(updated_tb_file_path) and is_file_open(updated_tb_file_path):
            messagebox.showerror("Error",
                                 f"Updated TB file is open: {os.path.basename(updated_tb_file_path)}. Please close it before proceeding.")
            return

        if is_manual.get():
            manual_function(tb_file_path.get(), pwp_file_path.get(), save_dir.get(), selected_promo.get())
        else:
            regular_function(tb_file_path.get(), pwp_file_path.get(), save_dir.get(), selected_promo.get())
        adjust_columns(save_dir.get(), updated_tb_file_path)

        # Reset the selections after processing
        tb_file_path.set("")
        pwp_file_path.set("")
        save_dir.set("")
        selected_promo.set("Required Field")
        tb_label.configure(text="Not Selected")
        pwp_label.configure(text="Not Selected")
        save_dir_label.configure(text="Not Selected")
        promo_dropdown.configure(state="disabled", values=[])
        promo_dropdown.set("Select Promo")

    def adjust_columns(save_dir, updated_tb_file_path):
        df = pd.read_excel(updated_tb_file_path, sheet_name=None)

        with pd.ExcelWriter(updated_tb_file_path, engine='openpyxl') as writer:
            for sheet_name, sheet_df in df.items():
                sheet_df.replace({pd.NA: '', pd.NaT: '', 'NAN': ''}, inplace=True)
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width

    tb_button = create_button(tab, "Select TB File", select_tb_file)
    tb_button.grid(row=0, column=0, padx=10, pady=10, sticky='w')
    tb_label = ctk.CTkLabel(tab, text="Not Selected", wraplength=window_width - 200)
    tb_label.grid(row=0, column=1, padx=10, pady=10, sticky='w', columnspan=2)

    pwp_button = create_button(tab, "Select PWP File", select_pwp_file)
    pwp_button.grid(row=1, column=0, padx=10, pady=10, sticky='w')
    pwp_label = ctk.CTkLabel(tab, text="Not Selected", wraplength=window_width - 200)
    pwp_label.grid(row=1, column=1, padx=10, pady=10, sticky='w', columnspan=2)

    save_dir_button = create_button(tab, "Select TB Location", select_save_dir)
    save_dir_button.grid(row=2, column=0, padx=10, pady=10, sticky='w')
    save_dir_label = ctk.CTkLabel(tab, text="Not Selected", wraplength=window_width - 200)
    save_dir_label.grid(row=2, column=1, padx=10, pady=10, sticky='w', columnspan=2)

    promo_label = ctk.CTkLabel(tab, text="Select Promo")
    promo_label.grid(row=3, column=0, padx=10, pady=10, sticky='w')
    promo_dropdown = ctk.CTkOptionMenu(tab, variable=selected_promo, state="disabled")
    promo_dropdown.grid(row=3, column=1, padx=10, pady=10, sticky='w', columnspan=2)

    manual_toggle = ctk.CTkCheckBox(tab, text="Manual Process", variable=is_manual)
    manual_toggle.grid(row=4, column=0, padx=10, pady=10, sticky='w')

    process_button = create_button(tab, "Process File", process_files)
    process_button.place(relx=1, rely=1, anchor='se')

    return tab


def process_files_generic(tb_file_path, pwp_file_path, save_dir, selected_promo, sheet_name):
    def find_closest_column(df, target_col):
        col_names = [str(col) for col in df.columns]
        matches = get_close_matches(target_col, col_names, n=1, cutoff=0.6)
        return matches[0] if matches else None

    updated_tb_file_path = os.path.join(save_dir, "Updated_" + os.path.basename(tb_file_path))

    tb_df = pd.read_excel(tb_file_path, sheet_name=0)
    tb_df_shifted = pd.concat([pd.DataFrame(columns=tb_df.columns), tb_df])
    tb_df_shifted.index = tb_df_shifted.index + 2
    tb_df_shifted.loc[1] = None
    tb_df_shifted = tb_df_shifted.sort_index()

    with pd.ExcelWriter(updated_tb_file_path, engine='openpyxl') as writer:
        tb_df_shifted.to_excel(writer, sheet_name="TB", index=False)

    tb_df = pd.read_excel(updated_tb_file_path, sheet_name=0, header=2)
    pwp_df = pd.read_excel(pwp_file_path, sheet_name=sheet_name, header=None)
    pwp_df.columns = pwp_df.iloc[5]
    pwp_df = pwp_df.drop(5)

    with pd.ExcelWriter(updated_tb_file_path, engine='openpyxl') as writer:
        tb_df_shifted.to_excel(writer, sheet_name="TB", index=False)
        pwp_df.to_excel(writer, sheet_name="PWP", index=False)

    messagebox.showinfo("Process Complete", f"Processing complete. Processing complete. Your TB file has been processed. {updated_tb_file_path}.")
    open_file_directory(updated_tb_file_path)


def lazada_process(tb_file_path, pwp_file_path, save_dir, selected_promo):
    import re

    def find_closest_column(df, target_col):
        col_names = [str(col) for col in df.columns]
        matches = get_close_matches(target_col, col_names, n=1, cutoff=0.6)
        return matches[0] if matches else None

    def clean_price(price_str):
        cleaned_str = re.sub(r'[^\d.]+', '', price_str)
        try:
            return float(cleaned_str)
        except ValueError:
            return None

    if not tb_file_path or not pwp_file_path or not save_dir or not selected_promo:
        messagebox.showerror("Error", "Please select all files, directories, and promo.")
        return

    updated_tb_file_path = os.path.join(save_dir, "Updated_" + os.path.basename(tb_file_path))

    if os.path.exists(updated_tb_file_path):
        os.chmod(updated_tb_file_path, 0o777)

    try:
        tb_df = pd.read_excel(tb_file_path, sheet_name=0, header=0)
    except PermissionError:
        messagebox.showerror("Error", f"Please close the TB file: {os.path.basename(tb_file_path)} first.")
        return
    except Exception as e:
        messagebox.showerror("Error", "Please select correct TB file")
        return

    try:
        pwp_df = pd.read_excel(pwp_file_path, sheet_name="Lzd | Campaign List", header=None)
    except PermissionError:
        messagebox.showerror("Error", f"Please close the PWP file: {os.path.basename(pwp_file_path)} first.")
        return

    pwp_df.columns = pwp_df.iloc[5]
    pwp_df = pwp_df.drop(5)

    shop_sku_tb_col = 'Shop Sku'
    shop_sku_pwp_col = 'SHOP SKU'
    campaign_price_col = 'Campaign Price'
    promo_name_col = 'Promo Name (Scheme)'
    discounted_price_col = 'Discounted Price/ASP (VATIN)'
    recommended_price_col = 'Recommended Price'

    shop_sku_tb_col = find_closest_column(tb_df, shop_sku_tb_col) or shop_sku_tb_col
    shop_sku_pwp_col = find_closest_column(pwp_df, shop_sku_pwp_col) or shop_sku_pwp_col
    campaign_price_col = find_closest_column(tb_df, campaign_price_col) or campaign_price_col
    promo_name_col = find_closest_column(pwp_df, promo_name_col) or promo_name_col
    discounted_price_col = find_closest_column(pwp_df, discounted_price_col) or discounted_price_col
    recommended_price_col = find_closest_column(tb_df, recommended_price_col) or recommended_price_col

    if not all([shop_sku_tb_col, shop_sku_pwp_col, campaign_price_col, promo_name_col, discounted_price_col, recommended_price_col]):
        messagebox.showerror("Error", "Required columns not found in the TB or PWP file.")
        return

    filtered_pwp_df = pwp_df[pwp_df[promo_name_col] == selected_promo]

    tb_df[shop_sku_tb_col] = tb_df[shop_sku_tb_col].astype(str).str.strip().str.upper()
    filtered_pwp_df.loc[:, shop_sku_pwp_col] = filtered_pwp_df[shop_sku_pwp_col].astype(str).str.strip().str.upper()

    good_for_upload_df = tb_df.copy()

    platform_df = pd.DataFrame(columns=[shop_sku_tb_col, campaign_price_col, 'Escalation Reason'])

    brand_df = pd.DataFrame(columns=[shop_sku_tb_col, 'Recommended Price'])

    tb_skus_in_pwp = set(filtered_pwp_df[shop_sku_pwp_col].dropna().unique())

    for idx, row in tb_df.iterrows():
        sku = row[shop_sku_tb_col]
        if sku in tb_skus_in_pwp:
            matching_row = filtered_pwp_df[filtered_pwp_df[shop_sku_pwp_col] == sku]
            if not matching_row.empty:
                discounted_price = matching_row.iloc[0][discounted_price_col]
                recommended_price = clean_price(str(row[recommended_price_col]))
                if isinstance(discounted_price, str):
                    discounted_price = clean_price(discounted_price)
                if recommended_price is not None and discounted_price is not None:
                    if recommended_price >= discounted_price:
                        good_for_upload_df.at[idx, campaign_price_col] = discounted_price
                    else:
                        platform_df = pd.concat([platform_df, pd.DataFrame(
                            {shop_sku_tb_col: [sku], campaign_price_col: [discounted_price], 'Escalation Reason': ['do not meet the reco price']})], ignore_index=True)
                else:
                    platform_df = pd.concat([platform_df, pd.DataFrame(
                        {shop_sku_tb_col: [sku], campaign_price_col: [discounted_price], 'Escalation Reason': ['invalid price format']})], ignore_index=True)
        else:
            brand_df = pd.concat([brand_df, pd.DataFrame({shop_sku_tb_col: [sku], 'Recommended Price': [row['Recommended Price']]})], ignore_index=True)

    good_for_upload_df = good_for_upload_df.dropna(subset=[campaign_price_col])

    pwp_skus_not_in_tb = filtered_pwp_df[~filtered_pwp_df[shop_sku_pwp_col].isin(tb_df[shop_sku_tb_col])]
    platform_df = pd.concat([platform_df, pwp_skus_not_in_tb[[shop_sku_pwp_col, discounted_price_col]].rename(
        columns={shop_sku_pwp_col: shop_sku_tb_col, discounted_price_col: campaign_price_col}).assign(
        **{'Escalation Reason': 'not eligible'})], ignore_index=True)

    brand_df.columns = ['Shop Sku', 'Recommended Price']

    try:
        with pd.ExcelWriter(updated_tb_file_path, engine='openpyxl') as writer:
            good_for_upload_df.replace({pd.NA: '', pd.NaT: '', 'NAN': ''}).to_excel(writer, sheet_name="Good for upload", index=False)
            platform_df.replace({pd.NA: '', pd.NaT: '', 'NAN': ''}).to_excel(writer, sheet_name="Platform", index=False)
            brand_df.replace({pd.NA: '', pd.NaT: '', 'NAN': ''}).to_excel(writer, sheet_name="Brand", index=False)

            for sheetname in writer.sheets:
                worksheet = writer.sheets[sheetname]
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width
        print(f"Processing complete. Your TB file has been processed. {updated_tb_file_path}.")
    except PermissionError as e:
        messagebox.showerror("Error", f"Please close the file: {updated_tb_file_path} first.")
        return

    messagebox.showinfo("Process Complete", f"Processing complete. Your TB file has been processed. {updated_tb_file_path}.")
    open_file_directory(updated_tb_file_path)


def lazada_manual_process(tb_file_path, pwp_file_path, save_dir, selected_promo):

    def find_closest_column(df, target_col):
        col_names = [str(col) for col in df.columns]
        matches = get_close_matches(target_col, col_names, n=1, cutoff=0.6)
        return matches[0] if matches else None

    if not tb_file_path or not pwp_file_path or not save_dir or not selected_promo:
        messagebox.showerror("Error", "Please select all files, directories, and promo.")
        return

    updated_tb_file_path = os.path.join(save_dir, "Updated_" + os.path.basename(tb_file_path))

    if os.path.exists(updated_tb_file_path):
        os.chmod(updated_tb_file_path, 0o777)

    try:
        tb_df = pd.read_excel(tb_file_path, sheet_name=0, header=0)
    except PermissionError:
        messagebox.showerror("Error", f"Please close the TB file: {os.path.basename(tb_file_path)} first.")
        return
    except Exception as e:
        messagebox.showerror("Error", "Please select correct TB file")
        return

    try:
        pwp_df = pd.read_excel(pwp_file_path, sheet_name="Lzd | Campaign List", header=None)
        pwp_df.columns = pwp_df.iloc[5]
        pwp_df = pwp_df.drop(5).reset_index(drop=True)
    except PermissionError:
        messagebox.showerror("Error", f"Please close the PWP file: {os.path.basename(pwp_file_path)} first.")
        return

    tb_sku_col = 'Shop SKU'
    pwp_sku_col = 'SHOP SKU'
    special_price_col = 'SpecialPrice'
    pwp_discounted_price_col = 'Discounted Price/ASP (VATIN)'
    special_price_start_col = 'SpecialPrice Start'
    special_price_end_col = 'SpecialPrice End'
    date_start_col = 'Date Start'
    date_end_col = 'Date End'
    time_start_col = 'Time Start'
    time_end_col = 'Time End'
    promo_name_col = 'Promo Name (Scheme)'

    tb_sku_col = find_closest_column(tb_df, tb_sku_col) or tb_sku_col
    pwp_sku_col = find_closest_column(pwp_df, pwp_sku_col) or pwp_sku_col
    special_price_col = find_closest_column(tb_df, special_price_col) or special_price_col
    pwp_discounted_price_col = find_closest_column(pwp_df, pwp_discounted_price_col) or pwp_discounted_price_col
    special_price_start_col = find_closest_column(tb_df, special_price_start_col) or special_price_start_col
    special_price_end_col = find_closest_column(tb_df, special_price_end_col) or special_price_end_col
    date_start_col = find_closest_column(pwp_df, date_start_col) or date_start_col
    date_end_col = find_closest_column(pwp_df, date_end_col) or date_end_col
    time_start_col = find_closest_column(pwp_df, time_start_col) or time_start_col
    time_end_col = find_closest_column(pwp_df, time_end_col) or time_end_col
    promo_name_col = find_closest_column(pwp_df, promo_name_col) or promo_name_col

    print(
        f"Identified Columns:\nShop Sku: {tb_sku_col}\nPWP Shop Sku: {pwp_sku_col}\nSpecial Price: {special_price_col}\nDiscounted Price: {pwp_discounted_price_col}\nDate Start: {date_start_col}\nDate End: {date_end_col}\nTime Start: {time_start_col}\nTime End: {time_end_col}\nPromo Name: {promo_name_col}")

    if promo_name_col not in pwp_df.columns:
        print(f"Error: '{promo_name_col}' column not found in PWP DataFrame.")
        return

    pwp_df[promo_name_col] = pwp_df[promo_name_col].astype(str).str.strip().str.lower()
    selected_promo = selected_promo.strip().lower()

    filtered_pwp_df = pwp_df[pwp_df[promo_name_col] == selected_promo].reset_index(drop=True)
    print("Filtered PWP DataFrame:")
    print(filtered_pwp_df.head())

    if filtered_pwp_df.empty:
        messagebox.showerror("Error", f"No data found for promo '{selected_promo}' in the PWP file.")
        return

    filtered_pwp_df[pwp_sku_col] = filtered_pwp_df[pwp_sku_col].astype(str).str.strip().str.upper()
    filtered_pwp_df[pwp_discounted_price_col] = pd.to_numeric(filtered_pwp_df[pwp_discounted_price_col],
                                                              errors='coerce')

    print("Normalized and Converted Data:")
    print(filtered_pwp_df[[pwp_sku_col, pwp_discounted_price_col]].head())

    updated_tb_df = tb_df[tb_df[tb_sku_col].isin(filtered_pwp_df[pwp_sku_col])].copy()

    if updated_tb_df.empty:
        messagebox.showerror("Error", "No matching SKUs found in TB data for the selected promo.")
        return

    # Clear the SpecialPrice Start and SpecialPrice End columns
    updated_tb_df[special_price_start_col] = ""
    updated_tb_df[special_price_end_col] = ""

    merged_df = pd.merge(updated_tb_df, filtered_pwp_df[
        [pwp_sku_col, pwp_discounted_price_col, date_start_col, date_end_col, time_start_col, time_end_col]],
                         left_on=tb_sku_col, right_on=pwp_sku_col, how='left')

    merged_df[special_price_col] = merged_df[pwp_discounted_price_col]

    # Combine the date from PWP with the time from PWP for start and end dates
    merged_df[special_price_start_col] = pd.to_datetime(merged_df[date_start_col].astype(str) + ' ' + filtered_pwp_df[time_start_col].astype(str))
    merged_df[special_price_end_col] = pd.to_datetime(merged_df[date_end_col].astype(str) + ' ' + filtered_pwp_df[time_end_col].astype(str))

    # Ensure the values are treated as text
    merged_df[special_price_start_col] = merged_df[special_price_start_col].dt.strftime('%Y-%m-%d %H:%M:%S')
    merged_df[special_price_end_col] = merged_df[special_price_end_col].dt.strftime('%Y-%m-%d %H:%M:%S')

    final_tb_df = merged_df[tb_df.columns]

    original_tb = pd.read_excel(tb_file_path, header=None)
    description_rows = original_tb.iloc[:4]

    final_df_with_description = pd.concat([description_rows, final_tb_df], ignore_index=True)

    print("Final DataFrame for upload:")
    print(final_df_with_description)

    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"

        # Write the description rows to the worksheet
        for r_idx, row in enumerate(dataframe_to_rows(description_rows, index=False, header=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)

        # Write the data starting from A5
        for r_idx, row in enumerate(dataframe_to_rows(final_tb_df, index=False, header=False), start=5):
            for c_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Ensure the SpecialPrice Start and SpecialPrice End columns are treated as text
        special_price_start_col_idx = final_tb_df.columns.get_loc(special_price_start_col) + 1
        special_price_end_col_idx = final_tb_df.columns.get_loc(special_price_end_col) + 1

        for cell in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=special_price_start_col_idx, max_col=special_price_start_col_idx):
            cell[0].number_format = '@'
        for cell in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=special_price_end_col_idx, max_col=special_price_end_col_idx):
            cell[0].number_format = '@'

        workbook.save(updated_tb_file_path)

        messagebox.showinfo("Process Complete", f"File has been updated and saved to {updated_tb_file_path}.")
    except PermissionError as e:
        print(f"PermissionError: {e}. Ensure the file is not open or read-only and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    open_file_directory(updated_tb_file_path)









def shopee_process(tb_file_path, pwp_file_path, save_dir, selected_promo):
    # Function to find the closest matching column name in the dataframe
    def find_closest_column(df, target_col):
        col_names = [str(col) for col in df.columns]
        matches = get_close_matches(target_col, col_names, n=1, cutoff=0.6)
        return matches[0] if matches else None

    # Check if any required input is missing
    if not tb_file_path or not pwp_file_path or not save_dir or not selected_promo or selected_promo == "Required Field":
        messagebox.showerror("Error", "Please select all files, directories, and promo.")
        return

    # Define the path for the updated TB file
    updated_tb_file_path = os.path.join(save_dir, "Updated_" + os.path.basename(tb_file_path))

    # Change file permissions if the updated file already exists
    if os.path.exists(updated_tb_file_path):
        os.chmod(updated_tb_file_path, 0o777)

    # Try to read the TB file
    try:
        tb_df = pd.read_excel(tb_file_path, sheet_name=0, header=0)
    except PermissionError:
        messagebox.showerror("Error", f"Please close the TB file: {os.path.basename(tb_file_path)} first.")
        return
    except Exception as e:
        messagebox.showerror("Error", "Please select correct TB file")
        return

    # Try to read the PWP file
    try:
        pwp_df = pd.read_excel(pwp_file_path, sheet_name="Shp | Campaign List", header=None)
    except PermissionError:
        messagebox.showerror("Error", f"Please close the PWP file: {os.path.basename(pwp_file_path)} first.")
        return

    # Set the column names for PWP dataframe
    pwp_df.columns = pwp_df.iloc[5]
    pwp_df = pwp_df.drop(5)

    # Find the closest matching column names
    tb_id_col = find_closest_column(tb_df, 'Variation ID')
    pwp_id_col = find_closest_column(pwp_df, 'Variation ID')
    campaign_price_col = find_closest_column(tb_df, 'Recommended Campaign Price')
    promo_name_col = find_closest_column(pwp_df, 'Promo Name (Scheme)')
    discounted_price_col = find_closest_column(pwp_df, 'Discounted Price/ASP (VATIN)')
    sales_price_col = find_closest_column(tb_df, 'Campaign Price')

    if not all([tb_id_col, pwp_id_col, campaign_price_col, promo_name_col, discounted_price_col, sales_price_col]):
        messagebox.showerror("Error", "Required columns not found in the TB or PWP file.")
        return

    # Check if the promo name column is found in PWP dataframe
    if promo_name_col not in pwp_df.columns:
        print(f"Error: '{promo_name_col}' column not found in PWP DataFrame.")
        return

    # Filter the PWP dataframe by the selected promo
    filtered_pwp_df = pwp_df[pwp_df[promo_name_col] == selected_promo]

    # Normalize and clean the SKU IDs for matching
    tb_df[tb_id_col] = tb_df[tb_id_col].astype(str).str.strip().str.upper()
    filtered_pwp_df.loc[:, pwp_id_col] = filtered_pwp_df[pwp_id_col].astype(str).str.strip().str.upper()

    # Convert price columns to numeric for comparison
    tb_df[campaign_price_col] = pd.to_numeric(tb_df[campaign_price_col], errors='coerce')
    filtered_pwp_df.loc[:, discounted_price_col] = pd.to_numeric(filtered_pwp_df[discounted_price_col], errors='coerce')

    # Create a copy of the TB dataframe to store updated prices
    good_for_upload_df = tb_df.copy()
    platform_df = pd.DataFrame(columns=[tb_id_col, campaign_price_col, 'Escalation Reason'])
    brand_df = pd.DataFrame(columns=[tb_id_col, campaign_price_col])

    # Create a set of unique SKUs from the filtered PWP dataframe
    tb_ids_in_pwp = set(filtered_pwp_df[pwp_id_col].dropna().unique())

    # Update the TB dataframe with prices from the PWP dataframe
    for idx, row in tb_df.iterrows():
        product_id = row[tb_id_col]
        if product_id in tb_ids_in_pwp:
            matching_rows = filtered_pwp_df[filtered_pwp_df[pwp_id_col] == product_id]
            if not matching_rows.empty:
                for _, matching_row in matching_rows.iterrows():
                    discounted_price = matching_row[discounted_price_col]
                    good_for_upload_df.at[idx, sales_price_col] = discounted_price
                    break
        else:
            brand_df = pd.concat([brand_df, pd.DataFrame({tb_id_col: [product_id], campaign_price_col: [row[campaign_price_col]]})], ignore_index=True)

    # Remove rows with missing campaign prices
    good_for_upload_df = good_for_upload_df.dropna(subset=[sales_price_col])

    # Find SKUs in PWP dataframe that are not in TB dataframe
    pwp_ids_not_in_tb = filtered_pwp_df[~filtered_pwp_df[pwp_id_col].isin(tb_df[tb_id_col])]
    platform_df = pd.concat([platform_df, pwp_ids_not_in_tb[[pwp_id_col, discounted_price_col]].rename(
        columns={pwp_id_col: tb_id_col, discounted_price_col: campaign_price_col}).assign(
        **{'Escalation Reason': 'not eligible'})], ignore_index=True)

    # Adjust the index of the good_for_upload_df for proper row insertion
    good_for_upload_df.index = good_for_upload_df.index + 3
    good_for_upload_df = good_for_upload_df.sort_index()

    # Save the updated dataframes to an Excel file
    try:
        with pd.ExcelWriter(updated_tb_file_path, engine='openpyxl') as writer:
            good_for_upload_df.to_excel(writer, sheet_name="Sheet1", index=False)
            platform_df.to_excel(writer, sheet_name="Platform", index=False)
            brand_df.to_excel(writer, sheet_name="Brand", index=False)
            for sheetname in writer.sheets:
                worksheet = writer.sheets[sheetname]
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width
        print(f"Processing complete. Your TB file has been processed. {updated_tb_file_path}.")
    except PermissionError as e:
        messagebox.showerror("Error", f"Please close the file: {updated_tb_file_path} first.")
        return

    # Notify the user that processing is complete and open the directory
    messagebox.showinfo("Process Complete", f"Processing complete. Your TB file has been processed. {updated_tb_file_path}.")
    open_file_directory(updated_tb_file_path)


def shopee_manual_process(tb_file_path, pwp_file_path, save_dir, selected_promo):
    def find_closest_column(df, target_col):
        col_names = [str(col) for col in df.columns]
        matches = get_close_matches(target_col, col_names, n=1, cutoff=0.6)
        return matches[0] if matches else None

    if not tb_file_path or not pwp_file_path or not save_dir or not selected_promo:
        messagebox.showerror("Error", "Please select all files, directories, and promo.")
        return

    updated_tb_file_path = os.path.join(save_dir, "Updated_" + os.path.basename(tb_file_path))

    if os.path.exists(updated_tb_file_path):
        os.chmod(updated_tb_file_path, 0o777)

    try:
        tb_df = pd.read_excel(tb_file_path, sheet_name=0, header=0)
    except PermissionError:
        messagebox.showerror("Error", f"Please close the TB file: {os.path.basename(tb_file_path)} first.")
        return
    except Exception as e:
        messagebox.showerror("Error", "Please select correct TB file")
        return

    try:
        pwp_df = pd.read_excel(pwp_file_path, sheet_name="Shp | Campaign List", header=None)
        pwp_df.columns = pwp_df.iloc[5]
        pwp_df = pwp_df.drop(5).reset_index(drop=True)
    except PermissionError:
        messagebox.showerror("Error", f"Please close the PWP file: {os.path.basename(pwp_file_path)} first.")
        return

    tb_id_col = 'Variation ID'
    pwp_id_col = 'Variation ID'
    tb_product_id_col = 'Product ID'
    pwp_product_id_col = 'Product ID'
    discount_price_col = 'Discount price'
    pwp_discounted_price_col = 'Discounted Price/ASP (VATIN)'
    promo_name_col = 'Promo Name (Scheme)'

    tb_id_col = find_closest_column(tb_df, tb_id_col) or tb_id_col
    pwp_id_col = find_closest_column(pwp_df, pwp_id_col) or pwp_id_col
    tb_product_id_col = find_closest_column(tb_df, tb_product_id_col) or tb_product_id_col
    pwp_product_id_col = find_closest_column(pwp_df, pwp_product_id_col) or pwp_product_id_col
    discount_price_col = find_closest_column(tb_df, discount_price_col) or discount_price_col
    pwp_discounted_price_col = find_closest_column(pwp_df, pwp_discounted_price_col) or pwp_discounted_price_col
    promo_name_col = find_closest_column(pwp_df, promo_name_col) or promo_name_col

    if not all([tb_id_col, pwp_id_col, tb_product_id_col, pwp_product_id_col, discount_price_col, pwp_discounted_price_col, promo_name_col]):
        messagebox.showerror("Error", "Required columns not found in the TB or PWP file.")
        return

    print(f"Identified Columns:\nVariation ID: {tb_id_col}\nPWP Variation ID: {pwp_id_col}\nProduct ID: {tb_product_id_col}\nPWP Product ID: {pwp_product_id_col}\nDiscount price: {discount_price_col}\nPromo Name: {promo_name_col}\nDiscounted Price: {pwp_discounted_price_col}")

    if promo_name_col not in pwp_df.columns:
        print(f"Error: '{promo_name_col}' column not found in PWP DataFrame.")
        return

    pwp_df[promo_name_col] = pwp_df[promo_name_col].astype(str).str.strip().str.lower()
    selected_promo = selected_promo.strip().lower()

    filtered_pwp_df = pwp_df[pwp_df[promo_name_col] == selected_promo].reset_index(drop=True)
    print("Filtered PWP DataFrame:")
    print(filtered_pwp_df.head())

    if filtered_pwp_df.empty:
        messagebox.showerror("Error", f"No data found for promo '{selected_promo}' in the PWP file.")
        return

    filtered_pwp_df[pwp_id_col] = filtered_pwp_df[pwp_id_col].astype(str).str.strip().str.upper()
    filtered_pwp_df[pwp_discounted_price_col] = pd.to_numeric(filtered_pwp_df[pwp_discounted_price_col], errors='coerce')

    print("Normalized and Converted Data:")
    print(filtered_pwp_df[[pwp_id_col, pwp_discounted_price_col, pwp_product_id_col]].head())

    good_for_upload_df = tb_df.copy()
    tb_ids_in_pwp = set(filtered_pwp_df[pwp_id_col].dropna().unique())
    print(f"IDs in PWP: {tb_ids_in_pwp}")

    for idx, row in filtered_pwp_df.iterrows():
        product_id = row[pwp_id_col]
        pwp_product_id = row[pwp_product_id_col]
        discounted_price = row[pwp_discounted_price_col]
        new_row = pd.DataFrame({
            tb_product_id_col: [pwp_product_id],
            tb_id_col: [product_id],
            discount_price_col: [discounted_price]
        })
        good_for_upload_df = pd.concat([good_for_upload_df, new_row], ignore_index=True)

    good_for_upload_df[tb_product_id_col] = good_for_upload_df[tb_product_id_col].astype(str)
    good_for_upload_df[tb_id_col] = good_for_upload_df[tb_id_col].astype(str)

    print("Good for upload DataFrame:")
    print(good_for_upload_df)

    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"



        for c_idx, value in enumerate(good_for_upload_df.columns, start=1):
            worksheet.cell(row=1, column=c_idx, value=value)

        for r_idx, row in enumerate(dataframe_to_rows(good_for_upload_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
            if column in ['B', 'C']:
                worksheet.column_dimensions[column].width = max(worksheet.column_dimensions[column].width, 15)

            if column in [tb_product_id_col, tb_id_col]:
                for cell in col:
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.number_format = '@'

        workbook.save(updated_tb_file_path)
        print(f"Processing complete. Your manual file has been processed. {updated_tb_file_path}.")
    except PermissionError as e:
        print(f"PermissionError: {e}. Ensure the file is not open or read-only and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    messagebox.showinfo("Process Complete", f"Processing complete. Your manual file has been processed. {updated_tb_file_path}.")
    open_file_directory(updated_tb_file_path)


def tiktok_manual_process(tb_file_path, pwp_file_path, save_dir, selected_promo):
    def find_closest_column(df, target_col):
        col_names = [str(col) for col in df.columns]
        matches = get_close_matches(target_col, col_names, n=1, cutoff=0.6)
        return matches[0] if matches else None

    if not tb_file_path or not pwp_file_path or not save_dir or not selected_promo:
        messagebox.showerror("Error", "Please select all files, directories, and promo.")
        return

    updated_tb_file_path = os.path.join(save_dir, "Updated_" + os.path.basename(tb_file_path))

    if os.path.exists(updated_tb_file_path):
        os.chmod(updated_tb_file_path, 0o777)

    tb_df = pd.read_excel(tb_file_path, sheet_name=0, header=0, dtype=str)
    if len(tb_df.columns) >= 2:
        tb_df = tb_df.iloc[:, :3]
        tb_df.columns = ["Product_id (required)", "SKU_id (required)", "Deal Price (required)"]
    else:
        messagebox.showerror("Error", "The TB file does not have the expected number of columns.")
        return

    print("TB DataFrame after loading with specified headers:")
    print(tb_df.head())

    original_tb_df = tb_df.copy()

    if tb_df.empty:
        print("TB DataFrame is empty after loading. Continuing to populate with PWP data.")

    pwp_df = pd.read_excel(pwp_file_path, sheet_name="TikTok | Campaign List", header=None, dtype=str)
    pwp_df.columns = pwp_df.iloc[5]
    pwp_df = pwp_df.drop(5).reset_index(drop=True)
    print("PWP DataFrame after setting headers:")
    print(pwp_df.head())

    tb_id_col = 'SKU_id (required)'
    pwp_id_col = 'SKU ID'
    tb_product_id_col = 'Product_id (required)'
    pwp_product_id_col = 'Product Id'
    campaign_price_col = 'Deal Price (required)'
    discounted_price_col = 'Discounted Price/ASP (VATIN)'
    promo_name_col = 'Promo Name (Scheme)'

    tb_id_col = find_closest_column(tb_df, tb_id_col) or tb_id_col
    pwp_id_col = find_closest_column(pwp_df, pwp_id_col) or pwp_id_col
    tb_product_id_col = find_closest_column(tb_df, tb_product_id_col) or tb_product_id_col
    pwp_product_id_col = find_closest_column(pwp_df, pwp_product_id_col) or pwp_product_id_col
    campaign_price_col = find_closest_column(tb_df, campaign_price_col) or campaign_price_col
    discounted_price_col = find_closest_column(pwp_df, discounted_price_col) or discounted_price_col
    promo_name_col = find_closest_column(pwp_df, promo_name_col) or promo_name_col

    if not all([tb_id_col, pwp_id_col, tb_product_id_col, pwp_product_id_col, campaign_price_col, discounted_price_col, promo_name_col]):
        messagebox.showerror("Error", "Required columns not found in the TB or PWP file.")
        return

    print(f"Identified Columns:\nSKU ID: {tb_id_col}\nPWP SKU ID: {pwp_id_col}\nProduct ID: {tb_product_id_col}\nPWP Product ID: {pwp_product_id_col}\nCampaign Price: {campaign_price_col}\nPromo Name: {promo_name_col}\nDiscounted Price: {discounted_price_col}")

    if promo_name_col not in pwp_df.columns:
        print(f"Error: '{promo_name_col}' column not found in PWP DataFrame.")
        return

    pwp_df[promo_name_col] = pwp_df[promo_name_col].astype(str).str.strip().str.lower()
    selected_promo = selected_promo.strip().lower()

    filtered_pwp_df = pwp_df[pwp_df[promo_name_col] == selected_promo].reset_index(drop=True)
    print("Filtered PWP DataFrame:")
    print(filtered_pwp_df.head())

    if filtered_pwp_df.empty:
        messagebox.showerror("Error", f"No data found for promo '{selected_promo}' in the PWP file.")
        return

    filtered_pwp_df[pwp_id_col] = filtered_pwp_df[pwp_id_col].astype(str).str.strip().str.upper()

    filtered_pwp_df[discounted_price_col] = pd.to_numeric(filtered_pwp_df[discounted_price_col], errors='coerce')

    print("Normalized and Converted Data:")
    print(filtered_pwp_df[[pwp_id_col, discounted_price_col, pwp_product_id_col]].head())

    good_for_upload_df = original_tb_df.copy()

    tb_ids_in_pwp = set(filtered_pwp_df[pwp_id_col].dropna().unique())
    print(f"IDs in PWP: {tb_ids_in_pwp}")

    for idx, row in filtered_pwp_df.iterrows():
        product_id = row[pwp_id_col]
        pwp_product_id = row[pwp_product_id_col]
        discounted_price = row[discounted_price_col]
        new_row = pd.DataFrame({
            tb_product_id_col: [pwp_product_id],
            tb_id_col: [product_id],
            campaign_price_col: [discounted_price]
        })
        good_for_upload_df = pd.concat([good_for_upload_df, new_row], ignore_index=True)

    good_for_upload_df[tb_product_id_col] = good_for_upload_df[tb_product_id_col].astype(str)
    good_for_upload_df[tb_id_col] = good_for_upload_df[tb_id_col].astype(str)

    print("Good for upload DataFrame:")
    print(good_for_upload_df)

    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Good for upload"

        # Load the TB workbook to get the header from D1 and E1
        tb_wb = load_workbook(tb_file_path)
        tb_ws = tb_wb.active
        header_d1 = tb_ws['D1'].value
        header_e1 = tb_ws['E1'].value

        # Set header and dimensions for columns D and E in the new worksheet
        worksheet['D2'] = header_d1
        worksheet['E2'] = header_e1

        for c_idx, value in enumerate(good_for_upload_df.columns, start=1):
            worksheet.cell(row=2, column=c_idx, value=value)

        for r_idx, row in enumerate(dataframe_to_rows(good_for_upload_df, index=False, header=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                cell.number_format = '00000' if isinstance(value, str) and value.isdigit() and len(value) > 10 else '@'

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
            if column in ['B', 'C']:
                worksheet.column_dimensions[column].width = max(worksheet.column_dimensions[column].width, 15)

        workbook.save(updated_tb_file_path)
        print(f"Processing complete. Your manual file has been processed. {updated_tb_file_path}.")
    except PermissionError as e:
        print(f"PermissionError: {e}. Ensure the file is not open or read-only and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    messagebox.showinfo("Process Complete", f"Processing complete. Your manual file has been processed. {updated_tb_file_path}.")
    open_file_directory(updated_tb_file_path)



def tiktok_process(tb_file_path, pwp_file_path, save_dir, selected_promo):
    def find_closest_column(df, target_col):
        col_names = [str(col) for col in df.columns]
        matches = get_close_matches(target_col, col_names, n=1, cutoff=0.6)
        return matches[0] if matches else None

    if not tb_file_path or not pwp_file_path or not save_dir or not selected_promo:
        messagebox.showerror("Error", "Please select all files, directories, and promo.")
        return

    updated_tb_file_path = os.path.join(save_dir, "Updated_" + os.path.basename(tb_file_path))

    if os.path.exists(updated_tb_file_path):
        os.chmod(updated_tb_file_path, 0o777)

    # Read TB file with header in the second row, assuming the file has headers but no data
    tb_df = pd.read_excel(tb_file_path, sheet_name=0, header=1)
    print("TB DataFrame after loading with header in the second row:")
    print(tb_df.head())

    # Make a copy of the original TB DataFrame
    original_tb_df = tb_df.copy()

    # Verify if TB DataFrame is empty
    if tb_df.empty:
        print("TB DataFrame is empty after loading. Continuing to populate with PWP data.")

    # Read PWP file
    pwp_df = pd.read_excel(pwp_file_path, sheet_name="TikTok | Campaign List", header=None)
    pwp_df.columns = pwp_df.iloc[5]
    pwp_df = pwp_df.drop(5).reset_index(drop=True)
    print("PWP DataFrame after setting headers:")
    print(pwp_df.head())

    # Column names identification
    tb_id_col = 'SKU ID'
    pwp_id_col = 'SKU ID'
    tb_product_id_col = 'Product ID'
    pwp_product_id_col = 'Product Id'
    campaign_price_col = 'Campaign price'
    discounted_price_col = 'Discounted Price/ASP (VATIN)'
    promo_name_col = 'Promo Name (Scheme)'

    tb_id_col = find_closest_column(tb_df, tb_id_col) or tb_id_col
    pwp_id_col = find_closest_column(pwp_df, pwp_id_col) or pwp_id_col
    tb_product_id_col = find_closest_column(tb_df, tb_product_id_col) or tb_product_id_col
    pwp_product_id_col = find_closest_column(pwp_df, pwp_product_id_col) or pwp_product_id_col
    campaign_price_col = find_closest_column(tb_df, campaign_price_col) or campaign_price_col
    discounted_price_col = find_closest_column(pwp_df, discounted_price_col) or discounted_price_col
    promo_name_col = find_closest_column(pwp_df, promo_name_col) or promo_name_col

    print(f"Identified Columns:\nSKU ID: {tb_id_col}\nPWP SKU ID: {pwp_id_col}\nProduct ID: {tb_product_id_col}\nPWP Product ID: {pwp_product_id_col}\nCampaign Price: {campaign_price_col}\nPromo Name: {promo_name_col}\nDiscounted Price: {discounted_price_col}")

    if promo_name_col not in pwp_df.columns:
        print(f"Error: '{promo_name_col}' column not found in PWP DataFrame.")
        return

    # Normalize promo names to ensure consistent comparison
    pwp_df[promo_name_col] = pwp_df[promo_name_col].astype(str).str.strip().str.lower()
    selected_promo = selected_promo.strip().lower()

    filtered_pwp_df = pwp_df[pwp_df[promo_name_col] == selected_promo].reset_index(drop=True)
    print("Filtered PWP DataFrame:")
    print(filtered_pwp_df.head())

    if filtered_pwp_df.empty:
        messagebox.showerror("Error", f"No data found for promo '{selected_promo}' in the PWP file.")
        return

    # Normalize IDs to ensure correct matching
    filtered_pwp_df[pwp_id_col] = filtered_pwp_df[pwp_id_col].astype(str).str.strip().str.upper()

    # Ensure numeric comparison for prices
    filtered_pwp_df[discounted_price_col] = pd.to_numeric(filtered_pwp_df[discounted_price_col], errors='coerce')

    print("Normalized and Converted Data:")
    print(filtered_pwp_df[[pwp_id_col, discounted_price_col, pwp_product_id_col]].head())

    # Create a new DataFrame for good_for_upload_df
    good_for_upload_df = original_tb_df.copy()

    tb_ids_in_pwp = set(filtered_pwp_df[pwp_id_col].dropna().unique())
    print(f"IDs in PWP: {tb_ids_in_pwp}")

    for idx, row in filtered_pwp_df.iterrows():
        product_id = row[pwp_id_col]
        pwp_product_id = row[pwp_product_id_col]
        discounted_price = row[discounted_price_col]
        new_row = pd.DataFrame({
            tb_product_id_col: [pwp_product_id],
            tb_id_col: [product_id],
            campaign_price_col: [discounted_price]
        })
        good_for_upload_df = pd.concat([good_for_upload_df, new_row], ignore_index=True)

    # Ensure the IDs are saved as text to prevent scientific notation
    good_for_upload_df[tb_product_id_col] = good_for_upload_df[tb_product_id_col].astype(str)
    good_for_upload_df[tb_id_col] = good_for_upload_df[tb_id_col].astype(str)

    print("Good for upload DataFrame:")
    print(good_for_upload_df)

    try:
        # Create a new Workbook and add the headers and data
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Good for upload"

        # Copy the description from A1 of the TB file
        tb_wb = load_workbook(tb_file_path)
        tb_ws = tb_wb.active
        description = tb_ws['A1'].value

        # Split the description into multiple lines
        description_lines = description.split('\n')
        description_wrapped = '\n'.join(description_lines)

        # Add description in A1
        worksheet['A1'] = description_wrapped
        worksheet.row_dimensions[1].height = 112.50
        worksheet.column_dimensions['A'].width = 52.73

        # Add headers in the second row
        for c_idx, value in enumerate(good_for_upload_df.columns, start=1):
            worksheet.cell(row=2, column=c_idx, value=value)

        # Write the updated DataFrame starting from the third row
        for r_idx, row in enumerate(dataframe_to_rows(good_for_upload_df, index=False, header=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                worksheet.cell(row=r_idx, column=c_idx, value=value)

        # Adjust column widths for readability
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
            if column in ['B', 'C']:  # Adjust specifically for columns B and C
                worksheet.column_dimensions[column].width = max(worksheet.column_dimensions[column].width, 15)

            if column in [tb_product_id_col, tb_id_col]:
                for cell in col:
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.number_format = '@'  # Setting text format to prevent scientific notation

        # Save the workbook
        workbook.save(updated_tb_file_path)
        print(f"Processing complete. Updated file saved to {updated_tb_file_path}.")
    except PermissionError as e:
        print(f"PermissionError: {e}. Ensure the file is not open or read-only and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    messagebox.showinfo("Process Complete", f"Processing complete. Updated file saved to {updated_tb_file_path}.")
    open_file_directory(updated_tb_file_path)


def main():
    app = ctk.CTk()
    app.title("GDEC Price Checker")

    window_width = 600
    window_height = 400

    screen_width = app.winfo_screenwidth()
    screen_height = app.winfo_screenheight()

    position_top = int(screen_height / 2 - window_height / 2)
    position_right = int(screen_width / 2 - window_width / 2)

    app.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")
    app.configure(fg_color='white')

    title = ctk.CTkLabel(app, text="Price Checker", font=('Arial', 40, 'bold'), text_color='black')
    title.pack(padx=20, pady=10)

    tabview = ctk.CTkTabview(app, width=window_width - 40, height=window_height - 100)
    tabview.pack(pady=20)

    create_tab(tabview, "Lazada", lazada_process, lazada_manual_process, "Lzd | Campaign List", window_width)
    create_tab(tabview, "Shopee", shopee_process, shopee_manual_process, "Shp | Campaign List", window_width)
    create_tab(tabview, "TikTok", tiktok_process, tiktok_manual_process, "TikTok | Campaign List", window_width)

    tabview.set("Lazada")

    version_label = ctk.CTkLabel(app, text="v2.2.4", font=('Arial', 15, 'italic'), text_color='grey')
    version_label.place(relx=1, rely=1, anchor='se', x=-10, y=7)

    app.mainloop()


if __name__ == "__main__":
    main()
